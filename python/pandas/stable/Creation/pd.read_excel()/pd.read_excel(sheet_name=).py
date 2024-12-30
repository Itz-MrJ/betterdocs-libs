import pandas as pd
from openpyxl import Workbook

# Create an Excel file with multiple sheets
file_path = "example.xlsx"

# Create a workbook and add data
wb = Workbook()
sheet1 = wb.active
sheet1.title = "Sheet1"
sheet1.append(["ID", "Name", "Age"])
sheet1.append([1, "Alice", 25])
sheet1.append([2, "Bob", 30])

sheet2 = wb.create_sheet(title="Sheet2")
sheet2.append(["ID", "City"])
sheet2.append([1, "New York"])
sheet2.append([2, "Los Angeles"])

sheet3 = wb.create_sheet(title="Sheet3")
sheet3.append(["ID", "Score"])
sheet3.append([1, 88.5])
sheet3.append([2, 92.0])

# Save the workbook
wb.save(file_path)

# Read the sheets
df = pd.read_excel(io=file_path, sheet_name=["Sheet2", "Sheet3"])
for name, df in df.items():
  print(df)
'''
Output:
   ID         City
0   1     New York
1   2  Los Angeles
   ID  Score
0   1   88.5
1   2   92.0
'''